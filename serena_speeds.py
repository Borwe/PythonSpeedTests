import speedtest
import os.path
import datetime
import openpyxl
from openpyxl.styles.borders import Border,Side
from openpyxl.styles import Font, PatternFill, Alignment

thin_border=Border(left=Side(style='thin'),
                                           right=Side(style='thin'),
                                           top=Side(style='thin'),
                                           bottom=Side(style='thin'))

def check_or_createFile():
    # check if file exists, if not the create it
    file=datetime.datetime.now().strftime("%B")
    file+='-'+datetime.datetime.now().strftime("%Y")+'-speedtests.xlsx'

    if os.path.isfile(file) == True:
        print("File ",file," exists")
    else:
        print("File doesn't exists, creating it")
        workbook=openpyxl.Workbook()

        #creating worksheets
        lqd_worksheet=workbook.active
        lqd_worksheet.title="LQD"
        jtl_worksheet=workbook.create_sheet('JTL')
        saf_worksheet=workbook.create_sheet('SAF')

        
        __fill_up_excel__(workbook,jtl_worksheet,'JTL')
        __fill_up_excel__(workbook,lqd_worksheet,'LQD')
        __fill_up_excel__(workbook,saf_worksheet,'SAF')
        
        workbook.save(file)
        workbook.close()

    return file

#fill up file with necessary titles for collumns
def __fill_up_excel__(workbook,worksheet,link_type):

    #add morning and afternoon text
    worksheet['B1']='Morning'
    worksheet['B1'].font=Font(bold=True,italic=True,underline='single')
    worksheet['B1'].fill=PatternFill(start_color='FFCC00',
                                     end_color='FFCC00', fill_type="solid")
    worksheet['O1']='Afternoon'
    worksheet['O1'].font=Font(bold=True,italic=True,underline='single')
    worksheet['O1'].fill=PatternFill(start_color='FFCC00',
                                     end_color='FFCC00', fill_type="solid")

    #create borders and titles
    worksheet['E2']=str(link_type+' LINK SPEEDTEST')
    worksheet['E2'].font=Font(bold=True,size=14)
    worksheet['E2'].alignment=Alignment(horizontal='center')
    worksheet.merge_cells('E2:H2')
    worksheet['R2']=str(link_type+' LINK SPEEDTEST')
    worksheet['R2'].font=Font(bold=True,size=14)
    worksheet['R2'].alignment=Alignment(horizontal='center')
    worksheet.merge_cells('R2:U2')

    #For dark borders
    dark_border = Border(left=Side(border_style='thick'),
                         right=Side(border_style='thick'),
                         top=Side(border_style='thick'),
                         bottom=Side(border_style='thick'))
    dark_border_font=Font(size=14,bold=True)
    dark_border_alignment=Alignment(horizontal='center')
    
    #morning table
    worksheet['C3']='UK'
    worksheet['C3'].alignment=dark_border_alignment
    worksheet['C3'].font=dark_border_font
    worksheet['C3'].border=dark_border
    worksheet['D3'].alignment=dark_border_alignment
    worksheet['D3'].font=dark_border_font
    worksheet['D3'].border=dark_border
    worksheet.merge_cells('C3:D3')

    worksheet['E3']='US'
    worksheet['E3'].alignment=dark_border_alignment
    worksheet['E3'].font=dark_border_font
    worksheet['E3'].border=dark_border
    worksheet['F3'].alignment=dark_border_alignment
    worksheet['F3'].font=dark_border_font
    worksheet['F3'].border=dark_border
    worksheet.merge_cells('E3:F3')

    worksheet['G3']='EUROPE'
    worksheet['G3'].alignment=dark_border_alignment
    worksheet['G3'].font=dark_border_font
    worksheet['G3'].border=dark_border
    worksheet['H3'].alignment=dark_border_alignment
    worksheet['H3'].font=dark_border_font
    worksheet['H3'].border=dark_border
    worksheet.merge_cells('G3:H3')
    
    worksheet['I3']='NAIROBI'
    worksheet['I3'].alignment=dark_border_alignment
    worksheet['I3'].font=dark_border_font
    worksheet['I3'].border=dark_border
    worksheet['J3'].alignment=dark_border_alignment
    worksheet['J3'].font=dark_border_font
    worksheet['J3'].border=dark_border
    worksheet.merge_cells('I3:J3')
    
    #DATE raw titles
    __fill_date_row_cells__(worksheet,'B4','DATE')
    __fill_date_row_cells__(worksheet,'C4','Download')
    __fill_date_row_cells__(worksheet,'D4','Upload')
    __fill_date_row_cells__(worksheet,'E4','Download')
    __fill_date_row_cells__(worksheet,'F4','Upload')
    __fill_date_row_cells__(worksheet,'G4','Download')
    __fill_date_row_cells__(worksheet,'H4','Upload')
    __fill_date_row_cells__(worksheet,'I4','Download')
    __fill_date_row_cells__(worksheet,'J4','Upload')
    __fill_date_row_cells__(worksheet,'K4','Remarks')
    __fill_date_row_cells__(worksheet,'L4','By')
    
    #evening table
    worksheet['P3']='UK'
    worksheet['P3'].alignment=dark_border_alignment
    worksheet['P3'].font=dark_border_font
    worksheet['P3'].border=dark_border
    worksheet['Q3'].alignment=dark_border_alignment
    worksheet['Q3'].font=dark_border_font
    worksheet['Q3'].border=dark_border
    worksheet.merge_cells('P3:Q3')

    worksheet['R3']='US'
    worksheet['R3'].alignment=dark_border_alignment
    worksheet['R3'].font=dark_border_font
    worksheet['R3'].border=dark_border
    worksheet['S3'].alignment=dark_border_alignment
    worksheet['S3'].font=dark_border_font
    worksheet['S3'].border=dark_border
    worksheet.merge_cells('R3:S3')

    worksheet['T3']='EUROPE'
    worksheet['T3'].alignment=dark_border_alignment
    worksheet['T3'].font=dark_border_font
    worksheet['T3'].border=dark_border
    worksheet['U3'].alignment=dark_border_alignment
    worksheet['U3'].font=dark_border_font
    worksheet['U3'].border=dark_border
    worksheet.merge_cells('T3:U3')

    worksheet['V3']='NAIROBI'
    worksheet['V3'].alignment=dark_border_alignment
    worksheet['V3'].font=dark_border_font
    worksheet['V3'].border=dark_border
    worksheet['W3'].alignment=dark_border_alignment
    worksheet['W3'].font=dark_border_font
    worksheet['W3'].border=dark_border
    worksheet.merge_cells('V3:W3')
    #DATE raw titles

    __fill_date_row_cells__(worksheet,'O4','DATE')
    __fill_date_row_cells__(worksheet,'P4','Download')
    __fill_date_row_cells__(worksheet,'Q4','Upload')
    __fill_date_row_cells__(worksheet,'R4','Download')
    __fill_date_row_cells__(worksheet,'S4','Upload')
    __fill_date_row_cells__(worksheet,'T4','Download')
    __fill_date_row_cells__(worksheet,'U4','Upload')
    __fill_date_row_cells__(worksheet,'V4','Download')
    __fill_date_row_cells__(worksheet,'W4','Upload')
    __fill_date_row_cells__(worksheet,'X4','Remarks')
    __fill_date_row_cells__(worksheet,'Y4','By')

def __fill_date_row_cells__(worksheet,cell,txt):
    #for date row cells
    date_row_font=Font(size=12, bold=True)
    dark_border_alignment=Alignment(horizontal='center')
    
    worksheet[cell]=txt
    worksheet[cell].border=thin_border
    worksheet[cell].font=date_row_font
    worksheet[cell].alignment=dark_border_alignment

class SerenaSpeedTester:

    def __init__(self):
        #hold Kenyan servers in list
        self.kenyan_servers=[]

        #hold United Kingdom servers in list
        self.uk_servers=[]
        
        #hold United States servers in list
        self.usa_servers=[]
        
        #Russia servers in list
        self.russia_servers=[]

        #speedtest object
        self.s=speedtest.Speedtest()

        self.__get_servers_based_on_our_four_region__()

        
    def __get_servers_based_on_our_four_region__(self):
        #get all speedtest.net servers
        self.__servers=self.s.get_servers()
        
        #display the ones in Kenya
        for point in self.__servers:
            self.server=self.__servers.get(point)
            #get country
            for part in self.server:
                # get servers that are in kenya
                if part.get('country').find('Kenya')!=-1:
                    self.kenyan_servers.append(part)
                     
                #get servers that are in UnitedKingdom
                if part.get('country').find('United Kingdom')!=-1:
                    self.uk_servers.append(part)

                #get servers that are in United States
                if part.get('country').find('United States')!=-1:
                    self.usa_servers.append(part)

                #get servers that are in United States
                if part.get('country').find('Russian Federation')!=-1:
                    self.russia_servers.append(part)

    #get list of IDs from server for usage
    def __get_country_servers_by_id__(self,country_servers):
        servers_by_id=[]
        for server in country_servers:
            servers_by_id.append(server.get('id'))

        return servers_by_id

    #turn bytes into Megabytes
    def __bytes_to_megabytes__(self,bytes):
        return round((bytes/(10**6)),2)

    #set time of checkup to evening
    def setTimeEvening(self):
        self.time='evening'

    #set time of checkup to morning
    def setTimeMorning(self):
        self.time='morning'

    #set worksheet to match JTL
    def setWorksheetJTL(self):
        self.worksheetName='JTL'

    #set worksheet to SAF
    def setWorkSheetSAF(self):
        self.worksheetName='SAF'

    #set worksheet to LQD
    def setWorkSheetLQD(self):
        self.worksheetName='LQD'

    #get current worksheet
    def getCurrentWorkSheet(self):
        return self.worksheetName

    #put data into file based on location in correct collumn
    def __enter_speeds_to_file__(self,file,download,upload,location,worksheet_name):
        
        #fake worksheet_name used (JTL)
        workbook=openpyxl.load_workbook(file)
        worksheet=workbook[worksheet_name]

        #USING FAKE LOCATION DEFAULT OF KENYA

        #get current date
        date=str(datetime.datetime.now().strftime("%m/%d/%Y"))

        #Store collumn letters
        date_collumns=['B','O']#[DAY,NIGHT]
        

        if(self.time=='evening'):
            cell_no=int(datetime.datetime.now().strftime("%d"))+5
            print('CELL: ',str(date_collumns[1]+'%d')%(cell_no))

            #set date cell
            worksheet[str(date_collumns[1]+'%d')%(cell_no)]=date
            worksheet[str(date_collumns[1]+'%d')%(cell_no)].border=thin_border

            if(location=='kenya'):
                kenya_evening=['V','W']

                #input download speeds
                worksheet[str(kenya_evening[0]+'%d')%(cell_no)]=str(download)
                worksheet[str(kenya_evening[0]+'%d')%(cell_no)].border=thin_border
                #input upload speeds
                worksheet[str(kenya_evening[1]+'%d')%(cell_no)]=str(upload)
                worksheet[str(kenya_evening[1]+'%d')%(cell_no)].border=thin_border

            if(location=='uk'):
                uk_evening=['P','Q']

                #input download speeds
                worksheet[str(uk_evening[0]+'%d')%(cell_no)]=str(download)
                worksheet[str(uk_evening[0]+'%d')%(cell_no)].border=thin_border
                #input upload speeds
                worksheet[str(uk_evening[1]+'%d')%(cell_no)]=str(upload)
                worksheet[str(uk_evening[1]+'%d')%(cell_no)].border=thin_border

            if(location=='usa'):
                usa_evening=['R','S']

                #input download speeds
                worksheet[str(usa_evening[0]+'%d')%(cell_no)]=str(download)
                worksheet[str(usa_evening[0]+'%d')%(cell_no)].border=thin_border
                #input upload speeds
                worksheet[str(usa_evening[1]+'%d')%(cell_no)]=str(upload)
                worksheet[str(usa_evening[1]+'%d')%(cell_no)].border=thin_border

            if(location=='russia'):
                russia_evening=['T','U']

                #input download speeds
                worksheet[str(russia_evening[0]+'%d')%(cell_no)]=str(download)
                worksheet[str(russia_evening[0]+'%d')%(cell_no)].border=thin_border
                #input upload speeds
                worksheet[str(russia_evening[1]+'%d')%(cell_no)]=str(upload)
                worksheet[str(russia_evening[1]+'%d')%(cell_no)].border=thin_border

        if(self.time=='morning'):
            cell_no=int(datetime.datetime.now().strftime("%d"))+5
            print('CELL: ',str(date_collumns[1]+'%d')%(cell_no))

            #set date cell
            worksheet[str(date_collumns[0]+'%d')%(cell_no)]=date
            worksheet[str(date_collumns[0]+'%d')%(cell_no)].border=thin_border

            if(location=='kenya'):
                kenya_evening=['I','J']

                #input download speeds
                worksheet[str(kenya_evening[0]+'%d')%(cell_no)]=str(download)
                worksheet[str(kenya_evening[0]+'%d')%(cell_no)].border=thin_border
                #input upload speeds
                worksheet[str(kenya_evening[1]+'%d')%(cell_no)]=str(upload)
                worksheet[str(kenya_evening[1]+'%d')%(cell_no)].border=thin_border

            if(location=='uk'):
                uk_evening=['C','D']

                #input download speeds
                worksheet[str(uk_evening[0]+'%d')%(cell_no)]=str(download)
                worksheet[str(uk_evening[0]+'%d')%(cell_no)].border=thin_border
                #input upload speeds
                worksheet[str(uk_evening[1]+'%d')%(cell_no)]=str(upload)
                worksheet[str(uk_evening[1]+'%d')%(cell_no)].border=thin_border

            if(location=='usa'):
                usa_evening=['E','F']

                #input download speeds
                worksheet[str(usa_evening[0]+'%d')%(cell_no)]=str(download)
                worksheet[str(usa_evening[0]+'%d')%(cell_no)].border=thin_border
                #input upload speeds
                worksheet[str(usa_evening[1]+'%d')%(cell_no)]=str(upload)
                worksheet[str(usa_evening[1]+'%d')%(cell_no)].border=thin_border

            if(location=='russia'):
                russia_evening=['G','H']

                #input download speeds
                worksheet[str(russia_evening[0]+'%d')%(cell_no)]=str(download)
                worksheet[str(russia_evening[0]+'%d')%(cell_no)].border=thin_border
                #input upload speeds
                worksheet[str(russia_evening[1]+'%d')%(cell_no)]=str(upload)
                worksheet[str(russia_evening[1]+'%d')%(cell_no)].border=thin_border
            
        workbook.save(file)
            

    #get speeds in Kenya
    def getSpeedsByInKenya(self):
        self.__kenya_server_ids=self.__get_country_servers_by_id__(self.kenyan_servers)
        self.s=speedtest.Speedtest()
        self.s.get_servers(self.__kenya_server_ids)
        self.s.get_best_server()
        self.s.download()
        self.s.upload()
        print(self.s.results.share())
        print("DOWNLOAD: ",self.__bytes_to_megabytes__(self.s.results.download))
        print("UPLOAD: ",self.__bytes_to_megabytes__(self.s.results.upload))

        file=check_or_createFile()
        print('adding download and uploads from Kenya to File: '+file)
        download=self.__bytes_to_megabytes__(self.s.results.download)
        upload=self.__bytes_to_megabytes__(self.s.results.upload)

        #fill up sheet(USING JTL AS TEST)
        self.__enter_speeds_to_file__(file,download,upload,'kenya',self.getCurrentWorkSheet())

    #get speeds in UK
    def getSpeedsByInUK(self):
        self.__uk_server_ids=self.__get_country_servers_by_id__(self.uk_servers)
        self.s=speedtest.Speedtest()
        self.s.get_servers(self.__uk_server_ids)
        self.s.get_best_server()
        self.s.download()
        self.s.upload()
        print(self.s.results.share())
        print("DOWNLOAD: ",self.__bytes_to_megabytes__(self.s.results.download))
        print("UPLOAD: ",self.__bytes_to_megabytes__(self.s.results.upload))

        file=check_or_createFile()
        print('adding download and uploads from UK to File: '+file)
        download=self.__bytes_to_megabytes__(self.s.results.download)
        upload=self.__bytes_to_megabytes__(self.s.results.upload)

        #fill up sheet(USING JTL AS TEST)
        self.__enter_speeds_to_file__(file,download,upload,'uk',self.getCurrentWorkSheet())

    #get speeds in USA
    def getSpeedsByInUS(self):
        self.__usa_server_ids=self.__get_country_servers_by_id__(self.usa_servers)
        self.s=speedtest.Speedtest()
        self.s.get_servers(self.__usa_server_ids)
        self.s.get_best_server()
        self.s.download()
        self.s.upload()
        print(self.s.results.share())
        print("DOWNLOAD: ",self.__bytes_to_megabytes__(self.s.results.download))
        print("UPLOAD: ",self.__bytes_to_megabytes__(self.s.results.upload))

        file=check_or_createFile()
        print('adding download and uploads from USA to File: '+file)
        download=self.__bytes_to_megabytes__(self.s.results.download)
        upload=self.__bytes_to_megabytes__(self.s.results.upload)

        #fill up sheet(USING JTL AS TEST)
        self.__enter_speeds_to_file__(file,download,upload,'usa',self.getCurrentWorkSheet())

    #get speeds in Russia
    def getSpeedsByInRussia(self):
        self.__russia_server_ids=self.__get_country_servers_by_id__(self.russia_servers)
        self.s=speedtest.Speedtest()
        self.s.get_servers(self.__russia_server_ids)
        self.s.get_best_server()
        self.s.download()
        self.s.upload()
        print(self.s.results.share())
        print("DOWNLOAD: ",self.__bytes_to_megabytes__(self.s.results.download))
        print("UPLOAD: ",self.__bytes_to_megabytes__(self.s.results.upload))

        file=check_or_createFile()
        print('adding download and uploads from Russia to File: '+file)
        download=self.__bytes_to_megabytes__(self.s.results.download)
        upload=self.__bytes_to_megabytes__(self.s.results.upload)

        #fill up sheet(USING JTL AS TEST)
        self.__enter_speeds_to_file__(file,download,upload,'russia',self.getCurrentWorkSheet())
